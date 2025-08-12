import streamlit as st
import matplotlib.pyplot as plt
import pandas as pd
import openpyxl
from io import BytesIO
from openpyxl.styles import PatternFill
import uuid
import io

# NEW: Plotly imports for interactive charts
import plotly.express as px
import plotly.graph_objects as go

# -------------------------------
# Configurations
# -------------------------------
valid_modules = ["SD", "FI", "CO", "HCM", "PM", "MM", "PP", "QM"]

# -------------------------------
# Mappings
# -------------------------------
code_to_designation = {
    "ZFIORI_CO": "Senior Manager",
    "ZFIORI_CO_COST_IN_ACCNT": "Senior Manager",
    "ZFIORI_CO_COST_OH_ACCNT": "Senior Manager",
    "ZFIORI_CO_COST_OH_ACCNT_T": "Senior Manager",
    "ZFIORI_CO_IS": "Senior Assistant Manager",
    "ZFIORI_HCM": "Senior Executive",
    "ZFIORI_MM_MDM": "Manager",
    "ZFIORI_MM_PO_RELEASE_MGR": "Senior Manager",
    "ZFIORI_MM_PR_RELEASE_DRT": "Senior Manager",
    "ZFIORI_MM_TEMP": "Senior Manager",
    "ZFIORI_PP": "Officer I",
    "ZFIORI_PPFI_MD_MNG": "Senior Executive",
    "ZFIORI_PPFI_PROD_EXE_MAN": "Senior Officer",
    "ZFIORI_PPFI_PROD_INVEN_MAN": "Senior Executive",
    "ZFIORI_PPFI_PROD_PLAN_MAN": "Senior Executive",
    "ZFIORI_PP_TEMP": "Senior Executive",
    "ZFIORI_SD": "Senior Manager",
    "Z_ABAP_DEVELOPER": "Executive",
    "Z_ABAP_USER_INTEG": "Senior Officer",
    "Z_AC_ABAP_ALL": "Senior Manager",
    "Z_CB_BANK_ENTRY": "Functional Manager",
    "Z_CB_BANK_ENTRY_1000": "Senior Assistant Manager",
    "Z_CB_BANK_ENTRY_2000": "Senior Manager",
    "Z_CB_BANK_ENTRY_3000": "Manager",
    "Z_CB_BANK_ENTRY_4000": "Manager",
    "Z_CB_BANK_ENTRY_5000": "Senior Assistant Manager",
    "Z_COMMON_ALL": "Senior Executive",
    "Z_CO_COMMON_REPORTS": "Officer II",
    "Z_CO_COMON_REPORT_IPAKCON_5000": "Senior Executive",
    "Z_CO_COST_FIORI_IN_ACCNT": "Senior Manager",
    "Z_CO_COST_FIORI_OH_ACCNT": "Senior Manager",
    "Z_SD_MD_CREATE_3000": "Senior Executive",
    "Z_SD_MD_CREATE_4000": "Executive",
    "Z_SD_MD_CREATE_50": "Senior Executive",
    "Z_SD_MD_CREATE_DS": "Officer I",
    "Z_SD_MD_CREATE_DS_10": "Senior Executive",
    "Z_SD_MD_DISPLAY": "Executive",
    "Z_SD_MD_DISPLAY_ALL": "Officer II",
    "Z_SD_REPORTS": "Senior Assistant Manager",
    "Z_SD_REPORTS_1000": "Manager",
    "Z_SD_REPORTS_2000": "Senior Assistant Manager",
    "Z_SD_REPORTS_ALL": "Senior Manager",
    "Z_SD_SALES_PERSON": "Senior Manager",
    "Z_SD_SALES_PERSON_NORTH": "Manager",
    "Z_SD_SALES_PERSON_SOUTH": "Senior Manager",
    "Z_SD_SUPPORT_USERS": "Senior Manager",
    "Z_SD_SUPPORT_USERS_4000": "Functional Manager",
    "Z_SD_SUPPORT_USERS_ALL": "Senior Manager",
    "Z_SD_SUPPORT_USERS_ALL_NORTH": "Senior Manager",
    "Z_SD_SUPPORT_USER_10": "Senior Manager",
    "Z_SD_SUPPORT_USER_30": "Senior Manager",
    "Z_SD_SUPPORT_USER_50": "Senior Manager",
    "Z_SD_USER_RSM": "Senior Manager",
    "Z_SD_USER_RSM_EXPORT": "Senior Assistant Manager",
    "Z_SD_USER_RSM_NORTH": "Senior Manager",
    "Z_SD_USER_RSM_SOUTH": "Senior Manager",
    "Z_SD_USER_RSM_SOUTH_COPY": "Manager",
    "Z_SD_USER_TL_NORTH": "Functional Manager",
    "Z_SD_USER_TL_SOUTH": "Senior Assistant Manager",
    "Z_SD_WASTESALE": "Senior Assistant Manager"
}

rank_based_roles = {
    "Senior Manager": [
        "ZFIORI_CO", "ZFIORI_CO_COST_IN_ACCNT", "ZFIORI_CO_COST_OH_ACCNT",
        "ZFIORI_CO_COST_OH_ACCNT_T", "ZFIORI_MM_PO_RELEASE_MGR", "ZFIORI_MM_PR_RELEASE_DRT",
        "ZFIORI_MM_TEMP", "ZFIORI_SD", "Z_AC_ABAP_ALL", "Z_CB_BANK_ENTRY_2000",
        "Z_CO_COST_FIORI_IN_ACCNT", "Z_CO_COST_FIORI_OH_ACCNT", "Z_SD_REPORTS_ALL",
        "Z_SD_SALES_PERSON", "Z_SD_SALES_PERSON_SOUTH", "Z_SD_SUPPORT_USERS",
        "Z_SD_SUPPORT_USERS_ALL", "Z_SD_SUPPORT_USERS_ALL_NORTH", "Z_SD_SUPPORT_USER_10",
        "Z_SD_SUPPORT_USER_30", "Z_SD_SUPPORT_USER_50", "Z_SD_USER_RSM",
        "Z_SD_USER_RSM_NORTH", "Z_SD_USER_RSM_SOUTH"
    ],
    "Senior Assistant Manager": [
        "ZFIORI_CO_IS", "Z_CB_BANK_ENTRY_1000", "Z_CB_BANK_ENTRY_5000", "Z_SD_REPORTS",
        "Z_SD_REPORTS_2000", "Z_SD_USER_RSM_EXPORT", "Z_SD_USER_TL_SOUTH",
        "Z_SD_WASTESALE"
    ],
    "Manager": [
        "ZFIORI_MM_MDM", "Z_CB_BANK_ENTRY_3000", "Z_CB_BANK_ENTRY_4000",
        "Z_SD_REPORTS_1000", "Z_SD_SALES_PERSON_NORTH", "Z_SD_USER_RSM_SOUTH_COPY"
    ],
    "Functional Manager": [
        "Z_CB_BANK_ENTRY", "Z_SD_SUPPORT_USERS_4000", "Z_SD_USER_TL_NORTH"
    ],
    "Senior Executive": [
        "ZFIORI_HCM", "ZFIORI_PPFI_MD_MNG", "ZFIORI_PPFI_PROD_INVEN_MAN",
        "ZFIORI_PPFI_PROD_PLAN_MAN", "ZFIORI_PP_TEMP", "Z_COMMON_ALL",
        "Z_CO_COMON_REPORT_IPAKCON_5000", "Z_SD_MD_CREATE_3000", "Z_SD_MD_CREATE_50",
        "Z_SD_MD_CREATE_DS_10"
    ],
    "Senior Officer": [
        "ZFIORI_PPFI_PROD_EXE_MAN", "Z_ABAP_USER_INTEG"
    ],
    "Officer I": [
        "ZFIORI_PP", "Z_SD_MD_CREATE_DS"
    ],
    "Officer II": [
        "Z_CO_COMMON_REPORTS", "Z_SD_MD_DISPLAY_ALL"
    ],
    "Executive": [
        "Z_ABAP_DEVELOPER", "Z_SD_MD_CREATE_4000", "Z_SD_MD_DISPLAY"
    ],
    "Junior Officer": [],
    "Grown": []
}

# -------------------------------
# Helper Functions (UNCHANGED LOGIC)
# -------------------------------
def get_expected_designation_from_role(role, code_to_designation):
    role = role.upper().strip() if isinstance(role, str) else ""
    if role in code_to_designation:
        return code_to_designation[role]
    for code, designation in code_to_designation.items():
        if role.startswith(code + "_") and role != code:
            return designation
    return None

def auto_detect_columns(df):
    role_col = next((col for col in df.columns if "role" in col.lower()), None)
    designation_col = next((col for col in df.columns if "designation" in col.lower()), None)
    return role_col, designation_col

def find_expected_roles(designation, rank_based_roles):
    designation = str(designation).strip().title() if isinstance(designation, str) else ""
    for title, roles in rank_based_roles.items():
        if title.title() == designation:
            return roles
    return []

def validate_file(df, role_col, designation_col, selected_module, code_to_designation, rank_based_roles):
    results = []
    df = df.fillna('')

    known_designations = set(v.title() for v in code_to_designation.values())

    for index, row in df.iterrows():
        designation_original = row[designation_col]
        role_original = row[role_col]

        designation = str(designation_original).strip().title() if pd.notnull(designation_original) else ""
        role = str(role_original).strip().upper() if pd.notnull(role_original) else ""

        role_is_empty = role == ""
        designation_is_empty = designation == ""
        is_designation_known = designation in known_designations

        if role_is_empty and designation_is_empty:
            results.append({
                "Module": selected_module,
                "Role": "",
                "Designation": "",
                "Valid": False,
                "Justification": "❌ Both Role and Designation fields are empty. This prevents validation and must be filled to ensure proper SAP access control.",
                "Suggested Role": "This entry is missing both fields. Please provide a valid role and designation."
            })
            continue

        if role_is_empty and not designation_is_empty:
            valid_roles = find_expected_roles(designation, rank_based_roles)
            suggested_role = valid_roles[0] if valid_roles else "please consult the SAP admin for role assignment"
            results.append({
                "Module": selected_module,
                "Role": "",
                "Designation": designation_original,
                "Valid": False,
                "Justification": f"❌ Designation '{designation_original}' provided, but Role is missing. Role is required for access control in {selected_module}.",
                "Suggested Role": f"Suggested role: '{suggested_role}' for designation '{designation_original}'."
            })
            continue

        expected_designation = get_expected_designation_from_role(role, code_to_designation)
        valid_roles_for_designation = find_expected_roles(designation, rank_based_roles)

        if expected_designation:
            if (expected_designation.title() == designation and
                role in valid_roles_for_designation):
                results.append({
                    "Module": selected_module,
                    "Role": role,
                    "Designation": designation_original,
                    "Valid": True,
                    "Justification": f"✅ Role '{role}' is valid for Designation '{designation_original}' in {selected_module}.",
                    "Suggested Role": "No changes needed."
                })
            else:
                suggested_role = valid_roles_for_designation[0] if valid_roles_for_designation else "please consult the SAP admin for correct role"
                results.append({
                    "Module": selected_module,
                    "Role": role,
                    "Designation": designation_original,
                    "Valid": False,
                    "Justification": (
                        f"❌ Role '{role}' maps to '{expected_designation}' in system, "
                        f"but used for '{designation_original}' or role is unauthorized."
                    ),
                    "Suggested Role": (
                        f"Replace '{role}' with '{suggested_role}' for '{designation_original}'."
                    )
                })
        else:
            if not is_designation_known:
                results.append({
                    "Module": selected_module,
                    "Role": role,
                    "Designation": designation_original,
                    "Valid": True,
                    "Justification": f"✅ External role-designation pair '{role} - {designation_original}' accepted as custom mapping.",
                    "Suggested Role": "No changes needed."
                })
            else:
                suggested_role = valid_roles_for_designation[0] if valid_roles_for_designation else "please consult the SAP admin for correct role"
                results.append({
                    "Module": selected_module,
                    "Role": role,
                    "Designation": designation_original,
                    "Valid": False,
                    "Justification": (
                        f"❌ Role '{role}' is unrecognized for known Designation '{designation_original}' in {selected_module}."
                    ),
                    "Suggested Role": (
                        f"Replace unrecognized role with '{suggested_role}' for '{designation_original}'."
                    )
                })

    return pd.DataFrame(results)

def highlight_invalid_rows(df):
    def style_func(row):
        styles = [''] * len(row)
        suggestion_col_idx = df.columns.get_loc("Suggested Role") if "Suggested Role" in df.columns else None
        if row['Valid'] == False:
            styles = ['background-color: #ffcccc'] * len(row)
        suggestion_text = str(row.get("Suggested Role", "")).strip().lower()
        if suggestion_col_idx is not None and suggestion_text != "no changes needed." and suggestion_text != "":
            styles[suggestion_col_idx] = 'background-color: #ccffff'
        return styles
    return df.style.apply(style_func, axis=1)

def download_excel(df):
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    df.to_excel(writer, sheet_name='Validation', index=False)
    writer.close()

    output.seek(0)
    wb = openpyxl.load_workbook(output)
    ws = wb.active

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    header = [cell.value for cell in ws[1]]
    valid_col_index = header.index("Valid") + 1

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        valid_cell = row[valid_col_index - 1]
        fill = green_fill if str(valid_cell.value).strip().lower() == "true" else red_fill
        for cell in row:
            cell.fill = fill

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# -------------------------------
# Helper function to render charts as resizable PNG (kept; not used with Plotly)
# -------------------------------
def render_resizable_chart(fig, chart_label, slider_key, default_width=300):
    with st.expander(f"📊 {chart_label}", expanded=True):
        width = st.slider(
            f"Adjust {chart_label} Size",
            min_value=200,
            max_value=800,
            value=default_width,
            step=50,
            key=slider_key
        )
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=120, bbox_inches="tight", facecolor="white")
        buf.seek(0)
        st.image(buf, width=width)

# -------------------------------
# Brand & Plotly colors (switched to a pro blue theme)
# -------------------------------
COLOR_ACCENT = "#2563eb"     # Blue-600
COLOR_ACCENT_DARK = "#1e3a8a"  # Indigo-800
COLOR_ACCENT_DEEP = "#0f172a"  # Slate-950
COLOR_VALID = "#16a34a"       # Green-600 (for success text only)
COLOR_INVALID = "#dc2626"     # Red-600

# Plotly color sequences (valid discrete colors)
COLOR_SEQ_MAIN = px.colors.qualitative.Vivid
COLOR_SEQ_ALT = px.colors.qualitative.Set2

PLOTLY_CONFIG = {
    "displaylogo": False,
    "modeBarButtonsToRemove": ["select2d", "lasso2d"],
    "toImageButtonOptions": {"format": "png", "filename": "chart", "height": 700, "width": 1200, "scale": 2}
}

# Background image (change URL if you want)
BACKGROUND_URL = "https://images.unsplash.com/photo-1556157382-97eda2d62296?auto=format&fit=crop&w=1600&q=80"

# -------------------------------
# Streamlit App
# -------------------------------
def run_app():
    st.set_page_config(page_title="SARKS", layout="wide")

    # Init state
    if 'validation_df' not in st.session_state:
        st.session_state.validation_df = None
    if 'page' not in st.session_state:
        st.session_state.page = "Upload & Validate"

    # -------------------- GLOBAL CSS (pro blue theme + background image) --------------------
    st.markdown(f"""
    <style>
        html, body, [class*="css"] {{
            font-family: "Inter", system-ui, -apple-system, Segoe UI, Roboto, Helvetica, Arial, sans-serif;
            -webkit-font-smoothing: antialiased;
            -moz-osx-font-smoothing: grayscale;
        }}

        /* App background image with a dark overlay for readability */
        .stApp {{
            background: linear-gradient(rgba(10, 14, 28, 0.65), rgba(10, 14, 28, 0.65)),
                        url('{BACKGROUND_URL}') no-repeat center center fixed;
            background-size: cover;
        }}

        /* Main container panels */
        .block-container {{
            padding-top: 1rem;
        }}

        /* Header */
        .hero {{
            background: linear-gradient(135deg, {COLOR_ACCENT_DEEP} 0%, {COLOR_ACCENT_DARK} 50%, {COLOR_ACCENT} 100%);
            padding: 16px 28px; border-radius: 14px; color: white;
            display: flex; align-items: center; justify-content: space-between;
            box-shadow: 0 8px 24px rgba(0,0,0,0.25); border: 1px solid rgba(255,255,255,0.08);
        }}
        .hero .title {{
            font-size: 26px; font-weight: 800; letter-spacing: 0.3px;
        }}
        .hero .subtitle {{
            font-size: 13px; opacity: 0.92; margin-top: 2px;
        }}

        /* KPI cards */
        .kpi {{
            background: rgba(255,255,255,0.9); border-radius: 14px; padding: 14px 16px;
            border: 1px solid #e6eaf0; box-shadow: 0 2px 12px rgba(0,0,0,0.06);
            backdrop-filter: blur(4px);
        }}
        .kpi .label {{ font-size: 12px; color: #6b7280; font-weight: 700; text-transform: uppercase; letter-spacing: .5px;}}
        .kpi .value {{ font-size: 24px; font-weight: 800; color: #0b152a; }}

        /* Sidebar */
        section[data-testid="stSidebar"] {{
            background: linear-gradient(180deg, rgba(11,16,33,0.95), rgba(15,23,54,0.95));
            border-right: 1px solid rgba(255,255,255,0.08);
        }}
        .sidebar-header {{
            color: #e0e7ff; text-align: center; padding: 14px 10px; border-radius: 12px;
            background: linear-gradient(135deg, {COLOR_ACCENT_DARK}, {COLOR_ACCENT});
            font-weight: 800; margin: 6px 0 14px 0; border: 1px solid rgba(255,255,255,0.15);
        }}
        .sidebar-tip {{
            color: #cbd5e1; font-size: 12px; margin: 8px 4px 14px 4px;
            background: rgba(30,58,138,0.25); padding: 8px 10px; border-radius: 10px; border: 1px solid rgba(255,255,255,0.08);
        }}
        /* Radio as navigation tabs */
        div[role="radiogroup"] > label {{
            background: transparent !important; border: 1px solid rgba(255,255,255,0.12) !important;
            border-radius: 12px; padding: 10px 12px !important; margin-bottom: 8px !important;
            color: #e2e8f0 !important; font-weight: 700;
        }}
        div[role="radiogroup"] > label:hover {{
            background: rgba(37, 99, 235, 0.12) !important; border-color: rgba(37, 99, 235, 0.45) !important;
        }}
        div[role="radiogroup"] > label[data-checked="true"] {{
            background: rgba(37, 99, 235, 0.25) !important; border-color: rgba(37, 99, 235, 0.65) !important;
            color: #ffffff !important;
        }}

        /* Footer */
        .custom-footer {{
            position: fixed; left: 0; bottom: 0; width: 100%;
            background: linear-gradient(135deg, {COLOR_ACCENT_DARK}, {COLOR_ACCENT});
            color: white; text-align: center; padding: 10px 8px;
            font-size: 13px; font-weight: 800; z-index: 9999;
            box-shadow: 0 -4px 14px rgba(0,0,0,0.35); border-top: 1px solid rgba(255,255,255,0.12);
        }}

        /* Dataframe header weight */
        .stDataFrame th {{ font-weight: 800 !important; }}
    </style>
    """, unsafe_allow_html=True)

    # -------------------- HEADER --------------------
    st.markdown(
        """
        <div class='hero'>
            <div>
                <div class='title'>📊 SARKS — Secure Access Role Knowledge System</div>
                <div class='subtitle'>Enterprise-grade validation of SAP roles vs. employee designations — with clear guidance & exportable insights.</div>
            </div>
        </div>
        """,
        unsafe_allow_html=True
    )

    # -------------------- SUB-HEADER --------------------
    st.markdown(
        """
        <p style='font-size:16px; color:#e5e7eb; margin-top:12px;'>
            🔍 Upload a file, run the validation, and explore interactive, drillable charts. Export results in one click.
        </p>
        """,
        unsafe_allow_html=True
    )

    # -------------------- SIDEBAR --------------------
    st.sidebar.markdown("<div class='sidebar-header'>🧭 Navigation</div>", unsafe_allow_html=True)
    st.sidebar.markdown(
        """
        <div class='sidebar-tip'>
            <b>Tip:</b> Validate first, then open <i>Validation Summary Charts</i> for a quick risk overview.
        </div>
        """, unsafe_allow_html=True
    )
    st.session_state.page = st.sidebar.radio(
        " ",  # hide visible label
        ["Upload & Validate", "Validation Summary Charts", "Download Results", "About"],
        index=["Upload & Validate", "Validation Summary Charts", "Download Results", "About"].index(st.session_state.page),
        key="page_selector"
    )
    if st.session_state.validation_df is not None:
        with st.sidebar.expander("📌 Quick Stats", expanded=True):
            try:
                total = len(st.session_state.validation_df)
                valid_count = int(st.session_state.validation_df["Valid"].sum())
                invalid_count = int((~st.session_state.validation_df["Valid"]).sum())
                st.write(f"• Total rows: **{total}**")
                st.write(f"• ✅ Valid: **{valid_count}**")
                st.write(f"• ❌ Invalid: **{invalid_count}**")
            except Exception:
                st.write("Stats will appear after validation.")
    st.sidebar.markdown("---")
    st.sidebar.caption("Theme: Indigo/Blue • Background image can be customized in code.")

    # -------------------- PAGE CONTENT --------------------
    if st.session_state.page == "Upload & Validate":
        st.markdown(
            """
            <p style='font-size:16px; color:#f3f4f6;'>
                Upload your <span style='color:#93c5fd;font-weight:700;'>.txt</span>, 
                <span style='color:#93c5fd;font-weight:700;'>.csv</span>, 
                <span style='color:#93c5fd;font-weight:700;'>.xls</span>, or 
                <span style='color:#93c5fd;font-weight:700;'>.xlsx</span> file to validate employee roles.
            </p>
            """,
            unsafe_allow_html=True
        )

        uploaded_file = st.file_uploader("📄 Upload File", type=["txt", "csv", "xls", "xlsx"])

        if uploaded_file:
            try:
                file_name = uploaded_file.name.lower()
                if file_name.endswith(".csv"):
                    df = pd.read_csv(uploaded_file)
                elif file_name.endswith(".txt"):
                    df = pd.read_csv(uploaded_file, delimiter="\t")
                elif file_name.endswith((".xls", ".xlsx")):
                    df = pd.read_excel(uploaded_file)
                else:
                    st.error("Unsupported file format.")
                    return
            except Exception as e:
                st.error(f"❌ Error reading file: {e}")
                return

            role_col, designation_col = auto_detect_columns(df)
            if not all([role_col, designation_col]):
                st.warning("⚠ Some columns not auto-detected. Please select manually:")
                role_col = st.selectbox("Select Role Column", df.columns, key="role_col_select")
                designation_col = st.selectbox("Select Designation Column", df.columns, key="designation_col_select")

            selected_module = st.selectbox("🧩 Select Module (Only one per file)", valid_modules)
            st.session_state.validation_df = validate_file(
                df, role_col, designation_col, selected_module,
                code_to_designation, rank_based_roles
            )
            st.success("✅ File validated! See results below or use the sidebar.")

            # KPI cards
            valdf = st.session_state.validation_df
            total = len(valdf)
            valid_count = int(valdf["Valid"].sum())
            invalid_count = int((~valdf["Valid"]).sum())
            invalid_pct = (invalid_count / total * 100) if total else 0

            c1, c2, c3, c4 = st.columns(4)
            with c1:
                st.markdown(f"<div class='kpi'><div class='label'>Total Rows</div><div class='value'>{total}</div></div>", unsafe_allow_html=True)
            with c2:
                st.markdown(f"<div class='kpi'><div class='label'>Valid</div><div class='value' style='color:{COLOR_VALID}'>{valid_count}</div></div>", unsafe_allow_html=True)
            with c3:
                st.markdown(f"<div class='kpi'><div class='label'>Invalid</div><div class='value' style='color:{COLOR_INVALID}'>{invalid_count}</div></div>", unsafe_allow_html=True)
            with c4:
                st.markdown(f"<div class='kpi'><div class='label'>Invalid %</div><div class='value'>{invalid_pct:.1f}%</div></div>", unsafe_allow_html=True)

            # Show validation results
            st.subheader("📋 Raw Validation Results")
            st.dataframe(highlight_invalid_rows(st.session_state.validation_df), use_container_width=True)

            errors_df = st.session_state.validation_df[st.session_state.validation_df["Valid"] == False]
            if not errors_df.empty:
                st.error("🚩 Detected mismatched roles! Suggested corrections are listed below:")
                st.dataframe(errors_df, use_container_width=True)

    elif st.session_state.page == "Validation Summary Charts":
        validation_df = st.session_state.get("validation_df", None)

        if validation_df is not None and not validation_df.empty:
            st.subheader("📈 Interactive Validation Summary")

            # Tabs for different charts
            tab1, tab2, tab3, tab4 = st.tabs([
                "✅ vs ❌ Overview",
                "Invalid by Module",
                "Invalid by Designation",
                "Treemap (Module → Designation)"
            ])

            # 1) Pie: Valid vs Invalid
            with tab1:
                vc = validation_df["Valid"].value_counts().rename_axis("Status").reset_index(name="Count")
                vc["Status"] = vc["Status"].map({True: "Valid", False: "Invalid"})
                color_map = {"Valid": "#22c55e", "Invalid": "#ef4444"}
                fig = px.pie(
                    vc,
                    names="Status",
                    values="Count",
                    color="Status",
                    color_discrete_map=color_map,
                    hole=0.35
                )
                fig.update_traces(
                    textposition='inside',
                    textinfo='percent+label',
                    hovertemplate="<b>%{label}</b><br>Count: %{value}<extra></extra>"
                )
                fig.update_layout(
                    legend_title="Status",
                    margin=dict(l=10, r=10, t=10, b=10),
                )
                st.plotly_chart(fig, use_container_width=True, config=PLOTLY_CONFIG)

            # 2) Bar: Invalid by Module (FIXED)
            with tab2:
                if "Module" in validation_df.columns:
                    invalid_by_module = (
                        validation_df[validation_df["Valid"] == False]["Module"]
                        .value_counts()
                        .rename_axis("Module")
                        .reset_index(name="Count")
                    )
                    if not invalid_by_module.empty:
                        fig = px.bar(
                            invalid_by_module,
                            x="Module", y="Count",
                            text="Count",
                            color="Module",
                            color_discrete_sequence=COLOR_SEQ_MAIN
                        )
                        fig.update_traces(
                            textposition='outside',
                            hovertemplate="<b>%{x}</b><br>Invalid: %{y}<extra></extra>"
                        )
                        fig.update_layout(
                            xaxis_title="Module", yaxis_title="Invalid Roles",
                            uniformtext_minsize=10, uniformtext_mode='show',
                            margin=dict(l=10, r=10, t=10, b=10),
                        )
                        st.plotly_chart(fig, use_container_width=True, config=PLOTLY_CONFIG)
                    else:
                        st.info("No invalid roles by module.")
                else:
                    st.info("Module column not found.")

            # 3) Bar: Invalid by Designation (Top N) (FIXED)
            with tab3:
                top_n = st.slider("Show Top N Designations", 5, 30, 15, 1, key="top_n_designations")
                if "Designation" in validation_df.columns:
                    invalid_designations = (
                        validation_df[validation_df["Valid"] == False]["Designation"]
                        .value_counts()
                        .rename_axis("Designation")
                        .reset_index(name="Count")
                        .head(top_n)
                    )
                    if not invalid_designations.empty:
                        fig = px.bar(
                            invalid_designations,
                            x="Designation", y="Count",
                            text="Count",
                            color="Designation",
                            color_discrete_sequence=COLOR_SEQ_ALT
                        )
                        fig.update_traces(
                            textposition='outside',
                            hovertemplate="<b>%{x}</b><br>Invalid: %{y}<extra></extra>"
                        )
                        fig.update_layout(
                            xaxis_title="Designation", yaxis_title="Invalid Count",
                            xaxis_tickangle=25,
                            margin=dict(l=10, r=10, t=10, b=10),
                        )
                        st.plotly_chart(fig, use_container_width=True, config=PLOTLY_CONFIG)
                    else:
                        st.info("No invalid roles by designation.")
                else:
                    st.info("Designation column not found.")

            # 4) Treemap: Module -> Designation for Invalid
            with tab4:
                if {"Module", "Designation"}.issubset(validation_df.columns):
                    invalid_subset = validation_df[validation_df["Valid"] == False]
                    if not invalid_subset.empty:
                        agg = (
                            invalid_subset.groupby(["Module", "Designation"])
                            .size()
                            .reset_index(name="Count")
                        )
                        fig = px.treemap(
                            agg,
                            path=["Module", "Designation"],
                            values="Count",
                            color="Module",
                            color_discrete_sequence=COLOR_SEQ_MAIN
                        )
                        fig.update_traces(
                            hovertemplate="<b>%{label}</b><br>Invalid: %{value}<extra></extra>"
                        )
                        fig.update_layout(margin=dict(l=10, r=10, t=10, b=10))
                        st.plotly_chart(fig, use_container_width=True, config=PLOTLY_CONFIG)
                    else:
                        st.info("No invalid records to visualize.")
                else:
                    st.info("Module or Designation column not found.")
        else:
            st.warning("📤 Please upload and validate a file first.")

    elif st.session_state.page == "Download Results":
        validation_df = st.session_state.get('validation_df', None)

        if validation_df is not None and not validation_df.empty:
            st.subheader("📥 Download Validated Results")
            st.markdown("""
            ✅ The downloaded Excel file will contain:
            - All uploaded roles and designations.
            - Validation status for each row (Valid/Invalid).
            - Suggested corrections if mismatches are found.
            - Highlighted invalid entries for easy review.
            """)
            st.download_button(
                label="📥 Download Excel",
                data=download_excel(validation_df),
                file_name="validated_roles.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.warning("📤 Please upload and validate a file first.")

    elif st.session_state.page == "About":
        st.markdown("### ℹ️ About SARKS")
        st.info("""
        **SARKS (Secure Access Role Knowledge System)** is a validation system designed to ensure that SAP roles 
        assigned to employees match their official designations.

        **How It Works:**
        - It scans uploaded Excel or CSV files containing roles and designations.
        - Matches each role code with predefined designation-to-role mappings.
        - Flags mismatches, missing roles/designations, and unauthorized access.
        - Suggests corrections and provides downloadable reports.

        ---
        ### 🔍 Validation Logic in SARKS
        1. **Both Role and Designation are Empty** — Flags incomplete records.
        2. **Role is Empty but Designation is Present** — Suggests a suitable role for the given designation.
        3. **Role Code Matches & Expected Designation Found** — Confirms valid or flags mismatched/unauthorized usage.
        4. **Role Code Not Found** — Differentiates custom external pairs from invalid roles for known designations.
        """)

    # -------------------- FOOTER --------------------
    st.markdown(
        """
        <div class="custom-footer">
            © 2025 • SARKS — Secure Access Role Knowledge System • Crafted by Sheikh Abdul Rehman
        </div>
        """,
        unsafe_allow_html=True
    )

if __name__ == "__main__":
    run_app()
